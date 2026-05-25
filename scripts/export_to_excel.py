# -*- coding: utf-8 -*-
"""导出做法库为 Excel 表格，便于查看和编辑。"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def export_methods_to_excel(output_path):
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    lib_path = os.path.join(base, "config", "methods_library.json")
    styles_path = os.path.join(base, "config", "table_styles.json")
    methods_lib = load_json(lib_path)
    styles = load_json(styles_path)

    section_labels = styles.get("section_labels", {})
    section_order = styles.get("section_order", [])

    wb = Workbook()

    # ===== 1. 做法清单 (主表) =====
    ws = wb.active
    ws.title = "做法清单"

    # 表头
    headers = ["序号", "建筑类型", "做法类型", "编号", "做法名称", "构造层次（从上到下）", "使用范围", "备注", "参考图集"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row = 2
    seq = 0
    # 按 section_order 顺序遍历
    for mtype in section_order:
        label = section_labels.get(mtype, mtype)
        for building_type in ["地下室", "住宅", "公建"]:
            methods = methods_lib.get(building_type, {}).get(mtype, [])
            for m in methods:
                seq += 1
                # 构造层次：合并为多行文本
                layers = m.get("layers", [])
                lines = []
                for layer in layers:
                    order = layer.get("order", "")
                    mat = layer.get("material", "")
                    thick = layer.get("thickness", "")
                    note = layer.get("note", "")
                    if thick and thick != "-":
                        thick_part = f"（{thick}）"
                    else:
                        thick_part = ""
                    note_part = f" [{note}]" if note else ""
                    lines.append(f"{order}. {mat}{thick_part}{note_part}")
                layers_text = "\n".join(lines)

                data = [
                    seq,                    # 序号
                    building_type,          # 建筑类型
                    label,                  # 做法类型
                    m.get("id", ""),        # 编号
                    m.get("name", ""),      # 做法名称
                    layers_text,            # 构造层次
                    m.get("usage", ""),     # 使用范围
                    m.get("notes", ""),     # 备注
                    m.get("reference", ""), # 参考图集
                ]
                for ci, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.border = thin_border
                    if ci in (1, 2, 3):
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = wrap_alignment
                    cell.font = Font(name="微软雅黑", size=9)
                row += 1

    # 列宽
    col_widths = [5, 8, 20, 14, 28, 45, 16, 20, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # ===== 2. 使用说明 =====
    ws2 = wb.create_sheet("使用说明")
    instructions = [
        ["修改说明", ""],
        ["", ""],
        ["如何修改做法", ""],
        ["1. 直接在「做法清单」工作表中修改对应单元格内容", ""],
        ["2. 构造层次：每个层次一行，格式为 序号. 材料名称（厚度） [备注]", ""],
        ["   示例：1. C20细石混凝土保护层（40） [配筋φ6@200双向]", ""],
        ["3. 使用范围：填写该做法适用于建筑的哪些部位/楼层", ""],
        ["   示例：适用于一般居室、起居室、书房等", ""],
        ["4. 备注：填写补充说明、参考图集号等", ""],
        ["", ""],
        ["如何新增做法", ""],
        ["在对应建筑类型和做法类型下方新增一行，填入完整信息即可", ""],
        ["编号规则：部位缩写-建筑类型-序号，如 WM-ZZ-003（屋面-住宅-003）", ""],
        ["", ""],
        ["修改完成后，请将 Excel 文件发回以便更新做法库", ""],
    ]
    for ri, (a, b) in enumerate(instructions, 1):
        cell_a = ws2.cell(row=ri, column=1, value=a)
        cell_b = ws2.cell(row=ri, column=2, value=b)
        if ri == 1:
            cell_a.font = Font(name="微软雅黑", size=12, bold=True)
        elif a.startswith("如何") or a.startswith("修改说明"):
            cell_a.font = Font(name="微软雅黑", size=11, bold=True)
        else:
            cell_a.font = Font(name="微软雅黑", size=10)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    # ===== 3. 构造层次参考 =====
    ws3 = wb.create_sheet("构造层次参考")
    ws3_headers = ["建筑类型", "做法类型", "编号", "做法名称", "层序", "材料", "厚度(mm)", "备注/说明"]
    for ci, h in enumerate(ws3_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row3 = 2
    for mtype in section_order:
        label = section_labels.get(mtype, mtype)
        for building_type in ["地下室", "住宅", "公建"]:
            methods = methods_lib.get(building_type, {}).get(mtype, [])
            for m in methods:
                layers = m.get("layers", [])
                if not layers:
                    # 空做法类型也显示一行
                    cell = ws3.cell(row=row3, column=1, value=building_type)
                    cell.border = thin_border
                    cell = ws3.cell(row=row3, column=2, value=label)
                    cell.border = thin_border
                    cell = ws3.cell(row=row3, column=3, value=m.get("id", ""))
                    cell.border = thin_border
                    cell = ws3.cell(row=row3, column=4, value=m.get("name", ""))
                    cell.border = thin_border
                    row3 += 1
                else:
                    for layer in layers:
                        data = [
                            building_type,
                            label,
                            m.get("id", ""),
                            m.get("name", ""),
                            layer.get("order", ""),
                            layer.get("material", ""),
                            layer.get("thickness", ""),
                            layer.get("note", ""),
                        ]
                        for ci, val in enumerate(data, 1):
                            cell = ws3.cell(row=row3, column=ci, value=val)
                            cell.border = thin_border
                            cell.font = Font(name="微软雅黑", size=9)
                            if ci in (1, 2):
                                cell.alignment = center_alignment
                            else:
                                cell.alignment = wrap_alignment
                        row3 += 1

    ws3_wide = [8, 20, 14, 28, 5, 40, 12, 30]
    for ci, w in enumerate(ws3_wide, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"已导出: {output_path}")
    return output_path


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "output", "建筑构造做法表_做法库.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    export_methods_to_excel(out)
